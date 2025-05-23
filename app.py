from fastapi import FastAPI, Depends, HTTPException, status, Query
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Optional
from openpyxl import Workbook, load_workbook
import os
from sqlalchemy import create_engine, Column, Integer, String, Boolean, ForeignKey
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship, Session

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allows all origins, for development only!
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class Task(BaseModel):
    description: str
    start_date: str
    deadline: str
    done: bool = False

TASKS_FILE = 'tasks.xlsx'

DATABASE_URL = 'sqlite:///./todo.db'
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base = declarative_base()

class User(Base):
    __tablename__ = 'users'
    id = Column(Integer, primary_key=True, index=True)
    username = Column(String, unique=True, index=True, nullable=False)
    password = Column(String, nullable=False)
    tasks = relationship('Task', back_populates='owner')

class Task(Base):
    __tablename__ = 'tasks'
    id = Column(Integer, primary_key=True, index=True)
    description = Column(String, nullable=False)
    start_date = Column(String, nullable=False)
    deadline = Column(String, nullable=False)
    done = Column(Boolean, default=False)
    owner_id = Column(Integer, ForeignKey('users.id'))
    owner = relationship('User', back_populates='tasks')

class UserCreate(BaseModel):
    username: str
    password: str

class UserLogin(BaseModel):
    username: str
    password: str

class TaskOut(BaseModel):
    id: int
    description: str
    start_date: str
    deadline: str
    done: bool
    class Config:
        from_attributes = True

Base.metadata.create_all(bind=engine)

def load_tasks_from_excel():
    loaded_tasks = []
    if os.path.exists(TASKS_FILE):
        wb = load_workbook(TASKS_FILE)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:
                done = bool(row[3]) if len(row) > 3 and row[3] is not None else False
                loaded_tasks.append(Task(description=row[0], start_date=row[1], deadline=row[2], done=done))
    return loaded_tasks

def save_tasks_to_excel(tasks):
    wb = Workbook()
    ws = wb.active
    ws.append(["description", "start_date", "deadline", "done"])
    for task in tasks:
        ws.append([task.description, task.start_date, task.deadline, task.done])
    wb.save(TASKS_FILE)

tasks: List[Task] = load_tasks_from_excel()

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@app.get("/tasks", response_model=List[TaskOut])
def get_tasks(user: Optional[str] = None, db: Session = Depends(get_db)):
    query = db.query(Task)
    if user:
        owner = db.query(User).filter(User.username == user).first()
        if owner:
            query = query.filter(Task.owner_id == owner.id)
        else:
            return []
    return query.all()

@app.post("/tasks", response_model=TaskOut)
def add_task(task: Task, user: str = Query(...), db: Session = Depends(get_db)):
    owner = db.query(User).filter(User.username == user).first()
    if not owner:
        raise HTTPException(status_code=400, detail="User not found")
    db_task = Task(
        description=task.description,
        start_date=task.start_date,
        deadline=task.deadline,
        done=task.done,
        owner_id=owner.id
    )
    db.add(db_task)
    db.commit()
    db.refresh(db_task)
    return db_task

@app.delete("/tasks/{task_idx}")
def delete_task(task_idx: int):
    if 0 <= task_idx < len(tasks):
        tasks.pop(task_idx)
        save_tasks_to_excel(tasks)
        return {"message": "Task removed"}
    return {"error": "Invalid index"}

@app.post('/register')
def register(user: UserCreate, db: Session = Depends(get_db)):
    existing_user = db.query(User).filter(User.username == user.username).first()
    if existing_user:
        raise HTTPException(status_code=400, detail='Username already registered')
    new_user = User(username=user.username, password=user.password)
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return {"message": "User registered successfully"}

@app.post('/login')
def login(user: UserLogin, db: Session = Depends(get_db)):
    db_user = db.query(User).filter(User.username == user.username, User.password == user.password).first()
    if not db_user:
        raise HTTPException(status_code=401, detail='Invalid username or password')
    # For simplicity, return username as token (not secure, just for demo)
    return {"token": db_user.username}